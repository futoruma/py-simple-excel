from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "Employee 1": {
        "apples": 34000,
        "melons": 58700,
        "oranges": 13800
    },
    "Employee 2": {
        "apples": 28000,
        "melons": 78200,
        "oranges": 31030
    },
    "Employee 3": {
        "apples": 45000,
        "melons": 11300,
        "oranges": 91520
    }
}

wb = Workbook()
ws = wb.active
ws.title = "Stats"

headings = ["Name"] + list(data["Employee 1"].keys())
ws.append(headings)

for employee in data:
    stats = list(data[employee].values())
    ws.append([employee] + stats)

for col in range(2, len(data["Employee 1"]) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, len(data["Employee 1"]) + 2):
    ws[get_column_letter(col) + "1"].font = Font(bold=True, color="0099CCFF")

wb.save("CalculatedStats.xlsx")